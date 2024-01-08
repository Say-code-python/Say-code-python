
""" first_name=input("What is your name? ")
is_online= False
print("Hello " + first_name)
birth_year=input("enter your birth year:")
age=2023 - int(birth_year)
print("Your age  " + first_name +" is  " +str(age)) """


"""first_number= float(input("enter first number "))
second_number= float(input("enter second number "))
sum= first_number + second_number
print(sum) """

"""course = "Python for beginners"
print(course.replace('for', '4'))
print(course.find('o'))
print('Python' in course)"""

"""print(10**2)
x=10
y=2
print(y!=x)"""

"""price=25
print(price>25 or price<30)
print( not price<30)"""

"""temperature = int(input("enter temperature: "))

if temperature >25:
    print("it's a sunny day")
elif temperature <=25:
    print("it's a nice day")"""

"""i=1
while i<=10:
    print(i* "*")
    i=i+1"""

"""names=["sayeeda", "naveed", "Sahil","Troy"]
names[0]="Sayeeda"
print(names[0:3])"""

#number=[1,2,3,4,5]
#number.insert(0,-1)
#numbers=range(5,11,2)
#for number in range(0,5):
  #  print(number)

#print("Sayeeda")
#print('o----')
#print(' ||||')
#print('*' * 10)

"""matrix= [
    [1,2,3],
    [4,5,6],
    [7,8,9]
]
matrix[0][2]=30
print(matrix[0])

for row in matrix:
    for item in row:
        print(item)

print(8 in matrix[2])"""

"""numbers=[1,2,3,5,6,7,1,3]
unique=[]
for number in numbers:
    if number not in unique:
        unique.append(number)
print( unique)

coordinates=[1, 2, 3]
x,y,z=coordinates
print(z)

customer={
    "name":input("enter your name: "),
"age":30,
    "is_verified":True
}
print(customer["name"])
customer["name"]="Sayeeda"
print(customer["name"])
print(customer.get("birth_day", "Nov 3 1971"))

phone=input("enter your choice phone number: ")
digits_declare = {
     "1": "one",
     "2": "two",
     "3": "three"
 }

output = ""

for chk in phone:
     output += digits_declare.get(chk, "!") + 
print(output)


#message = input(">")
def display_emoji(sentence):
    words = sentence.split(" ")
    emojis = {
        ":(" : "\U0001F601",
        ":)" : "\U0001F600"
        }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return(output)


message=input(">")
print(display_emoji(message))

# print(output + "\U0001F601")

def greet_msg(first_name,last_name):go out
    print(f'hello {first_name} . {last_name}!')
    print('welcome to MHA')


print('start')
greet_msg('Sayeeda', "Sultana")
greet_msg('Naveed', "MAS")
print('Finish')

def square(length):
    return length * length

print(square(3))

try:
    age= int(input('Age: '))
    income=80000
    risk=(income/age)/100
    print(risk)
    print(age)
except ValueError:
     print("Invalid input")
except ZeroDivisionError:
     print("Found Zero divided") 

class PointHim:
    def __init__(self,x,y):
        self.x = x
        self.y = y
    def move(self):
        print("Move")

    def draw(self):
        print("Draw")


point1= PointHim(20,30)
point1.move()

print(point1.x )
print(point1.y) 

class Person:
    def __init__(self,name):
        self.name=name
    def talk(self):
        print(f'Hi , I am {self.name}')

person1=Person('sayeeda')
person1.talk()
print(person1.name)

person2=Person('Naveed')
person2.talk()
print(person2.name) 

class Mammal:
    def walk(self):
        print("walk")

class dog(Mammal):
    def bark(self):
        print("bark")

class cat(Mammal):
    pass

dog1 = dog()
dog1.walk()
dog1.bark() 

cat1 = cat()
cat1.walk() 

import converters
from converters import kg_to_lbs
print(kg_to_lbs(80) ) 

import maxfinder
from maxfinder import max_find
numbers=[2,60,3,8,1]
max = max_find(numbers)
print(max) 

#from ecommerce.shipping import calc_shipping
from ecommerce import shipping

shipping.calc_shipping()
shipping.calc_shipping()

#ecommerce.shipping.calc_shipping() 

import random

for i in range(5):
    print(random.randint(10, 20))

leader_list=['Sayeeda','Naveed','Troy','Sahil']
print(random.choice(leader_list))

import random

class Dice:
    def roll(self):
        first_number = random.randint(1,6)
        second_number = random.randint(1,6)
        return first_number , second_number

dice1 = Dice()
print(dice1.roll())  


from pathlib import Path

path1 = Path()

for file in path1.glob('*'):
    print(file) """

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']
#cell = sheet['a1']
cell = sheet.cell(1,1)
#print(cell.value)
#print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell=sheet.cell(row,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')


